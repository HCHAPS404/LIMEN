"""PHASE 5C.2 official dataset unit tests (no LLM, no real dataset)."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from evals.llm.official_benchmark import case_cache_key, dataset_fingerprint_digest
from evals.llm.official_dataset import (
    canonical_resolution_candidates,
    discover_official_dataset,
    firewall_prompt,
)
from evals.llm.official_labels import normalize_official_label
from evals.llm.official_load import (
    RESULT_SHEET,
    load_official_tables,
    resolve_official_paths,
)
from evals.llm.official_metrics import (
    compute_degradation,
    compute_official_metrics,
    score_official_prediction,
    summarize_official_results,
)
from evals.llm.official_reconstruct import reconstruct_official_dataset
from evals.llm.prompts import official_advisory_user_prompt
from evals.llm.scorecard import recommend_primary_fallback


def _write_mini_official_xlsx(root: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = RESULT_SHEET
    ws.append(
        [
            "dialogo_id",
            "caso_id",
            "paciente_id",
            "dia_postop",
            "turno_idx",
            "hablante",
            "texto",
            "label_ground_truth",
            "estilo_paciente",
            "modelo_paciente",
            "modelo_agente",
            "capa",
            "generado_ts",
        ]
    )
    rows = [
        (1, "caso_tray_a", "pac_1", 1, 1, "paciente", "me duele", "verde", "x", "mp", "ma", "capa1_limpia", "ts"),
        (2, "caso_tray_a", "pac_1", 1, 2, "agente", "cuénteme más", "verde", "x", "mp", "ma", "capa1_limpia", "ts"),
        (3, "caso_tray_a", "pac_1", 1, 1, "paciente", "me duele mucho", "verde", "x", "mp", "ma", "capa2_ruidosa", "ts"),
        (4, "caso_tray_b", "pac_2", 2, 1, "paciente", "tengo fiebre", "rojo", "x", "mp", "ma", "capa1_limpia", "ts"),
        (5, "caso_tray_b", "pac_2", 2, 1, "paciente", "tengo fiebre alta", "rojo", "x", "mp", "ma", "capa2_ruidosa", "ts"),
    ]
    for row in rows:
        ws.append(list(row))
    wb.save(root / "dataset_final.xlsx")
    wb.close()

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = RESULT_SHEET
    ws2.append(
        [
            "trayectoria_id",
            "paciente_id",
            "dia_postop",
            "arquetipo_trayectoria",
            "dolor_nrs",
            "fiebre_c",
            "movilidad",
            "herida",
            "apetito",
            "sueno",
            "seed",
            "generado_ts",
        ]
    )
    ws2.append(["tray_a", "pac_1", 1, "stable", 2, 36.5, "ok", "ok", "ok", "ok", 1, "ts"])
    ws2.append(["tray_b", "pac_2", 2, "fever", 5, 39.0, "low", "bad", "low", "bad", 2, "ts"])
    wb2.save(root / "trayectorias_postop_silver.xlsx")
    wb2.close()

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = RESULT_SHEET
    ws3.append(
        [
            "paciente_id",
            "bundle_id",
            "synthea_runtime",
            "modulo_synthea",
            "procedimiento",
            "fecha_cirugia",
            "edad",
            "genero",
            "comorbilidades",
            "complicacion_encounter",
            "generado_ts",
        ]
    )
    ws3.append(["pac_1", "b1", "rt", "mod", "colecistectomía", "2026-01-01", 45, "F", "HTA", "none", "ts"])
    ws3.append(["pac_2", "b2", "rt", "mod", "apendicectomía", "2026-01-02", 30, "M", "", "fever", "ts"])
    wb3.save(root / "perfiles_clinicos_pacientes_silver_contest.xlsx")
    wb3.close()

    wb4 = Workbook()
    ws4 = wb4.active
    ws4.title = RESULT_SHEET
    ws4.append(
        [
            "paciente_id",
            "nombre_completo",
            "direccion",
            "ciudad",
            "departamento",
            "documento_cc",
            "eps",
            "source_country",
            "adapted_country",
            "adaptation_fields",
            "adaptation_ts",
        ]
    )
    ws4.append(["pac_1", "Ana", "calle", "BOG", "CUN", "123", "eps", "CO", "CO", "x", "ts"])
    wb4.save(root / "perfiles_pacientes_co.xlsx")
    wb4.close()


def test_canonical_resolution_order(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("LIMEN_DATASET_PATH", raising=False)
    labels = [label for label, _ in canonical_resolution_candidates()]
    assert labels[0] == "./dataset/"
    assert "./data/challenge/" in labels

    monkeypatch.setenv("LIMEN_DATASET_PATH", "/tmp/dataset")
    labels_with_env = [label for label, _ in canonical_resolution_candidates()]
    assert labels_with_env[0] == "LIMEN_DATASET_PATH"
    assert "./dataset/" in labels_with_env
    assert "./data/challenge/" in labels_with_env


def test_label_normalization() -> None:
    assert normalize_official_label("verde") == "GREEN"
    assert normalize_official_label("amarillo") == "YELLOW"
    assert normalize_official_label("rojo") == "RED"
    with pytest.raises(ValueError):
        normalize_official_label("purple")


def test_load_and_reconstruct_mini_dataset(tmp_path: Path) -> None:
    _write_mini_official_xlsx(tmp_path)
    paths = resolve_official_paths(tmp_path)
    assert paths is not None
    tables = load_official_tables(paths)
    reconstructed = reconstruct_official_dataset(tables)
    assert len(reconstructed.conversations) == 4
    assert reconstructed.stats["unique_case_ids"] == 2
    assert reconstructed.stats["layer_counts"]["capa1_limpia"] == 2
    assert reconstructed.stats["layer_counts"]["capa2_ruidosa"] == 2
    conv = reconstructed.conversations[0]
    assert all(set(t.keys()) <= {"turno_idx", "hablante", "texto"} for t in conv.ordered_turns)
    assert "complicacion_encounter" not in conv.known_clinical_profile
    assert "procedimiento" in conv.known_clinical_profile


def test_firewall_blocks_trajectory_and_labels() -> None:
    with pytest.raises(AssertionError):
        firewall_prompt("label_ground_truth=rojo conversación")
    with pytest.raises(AssertionError):
        firewall_prompt("trayectoria con dolor_nrs=8")
    prompt = official_advisory_user_prompt("paciente: hola", {"edad": 40})
    assert firewall_prompt(prompt) == prompt


def test_official_metrics_red_fn_and_orange() -> None:
    rows = [
        {"ground_truth": "RED", "predicted": "YELLOW", "exact_match": False, "red_false_negative": True},
        {"ground_truth": "RED", "predicted": "RED", "exact_match": True, "red_false_negative": False},
        {"ground_truth": "GREEN", "predicted": "ORANGE", "exact_match": False, "red_false_negative": False},
    ]
    metrics = compute_official_metrics(rows)
    assert metrics["red_false_negatives"] == 1
    assert metrics["red_total"] == 2
    assert metrics["accuracy"] == pytest.approx(1 / 3)

    scored_orange = score_official_prediction(
        ground_truth="YELLOW", predicted="ORANGE", valid_json=True
    )
    assert scored_orange["exact_match"] is False
    assert scored_orange["red_false_negative"] is False


def test_macro_f1_and_degradation() -> None:
    clean_rows = [
        {"ground_truth": "GREEN", "predicted": "GREEN", "exact_match": True, "layer": "capa1_limpia"},
        {"ground_truth": "RED", "predicted": "RED", "exact_match": True, "layer": "capa1_limpia"},
    ]
    noisy_rows = [
        {"ground_truth": "GREEN", "predicted": "YELLOW", "exact_match": False, "layer": "capa2_ruidosa"},
        {"ground_truth": "RED", "predicted": "YELLOW", "exact_match": False, "red_false_negative": True, "layer": "capa2_ruidosa"},
    ]
    summary = summarize_official_results(clean_rows + noisy_rows)
    assert summary["overall"]["n"] == 4
    deg = compute_degradation(summary["clean"], summary["noisy"])
    assert deg["accuracy_delta_clean_minus_noisy"] == pytest.approx(1.0)


def test_cache_key_stable() -> None:
    key_a = case_cache_key(
        case_id="caso_x",
        layer="capa1_limpia",
        commit_sha="abc",
        dataset_sha="def",
        model_digest="ghi",
        temperature=0.2,
        max_tokens=256,
    )
    key_b = case_cache_key(
        case_id="caso_x",
        layer="capa1_limpia",
        commit_sha="abc",
        dataset_sha="def",
        model_digest="ghi",
        temperature=0.2,
        max_tokens=256,
    )
    assert key_a == key_b
    key_c = case_cache_key(
        case_id="caso_x",
        layer="capa2_ruidosa",
        commit_sha="abc",
        dataset_sha="def",
        model_digest="ghi",
        temperature=0.2,
        max_tokens=256,
    )
    assert key_a != key_c


def test_fingerprint_digest_changes_with_content(tmp_path: Path) -> None:
    _write_mini_official_xlsx(tmp_path)
    paths = resolve_official_paths(tmp_path)
    assert paths is not None
    from evals.llm.official_dataset import fingerprint_dataset_file

    fps = [
        fingerprint_dataset_file(paths.dataset_final).to_dict(),
        fingerprint_dataset_file(paths.trayectorias).to_dict(),
    ]
    digest_a = dataset_fingerprint_digest(fps)
    digest_b = dataset_fingerprint_digest(fps)
    assert digest_a == digest_b


def test_scorecard_definitive_only_with_official_complete() -> None:
    base = {
        "status": "measured",
        "resolved_tag": "llama3.2:3b",
        "critical_safety_failures": 0,
        "safety_decision_contradiction_rate": 0.0,
        "unsupported_claim_rate": 0.0,
        "structured_output": {"schema_valid_rate": 1.0},
        "injection": {"overall_resist_rate": 1.0},
        "measured_case_count": 320,
        "official_red_false_negatives": 1,
        "official_macro_f1": 0.7,
    }
    results = [
        {**base, "model_id": "llama3.2:3b"},
        {**base, "model_id": "llama3.2:1b", "official_red_false_negatives": 2, "official_macro_f1": 0.5},
    ]
    provisional = recommend_primary_fallback(results, official_red_available=True, official_eval_complete=False)
    assert provisional["STATUS"] == "PROVISIONAL"
    definitive = recommend_primary_fallback(results, official_red_available=True, official_eval_complete=True)
    assert definitive["STATUS"] == "DEFINITIVE"


def test_discover_with_env(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _write_mini_official_xlsx(tmp_path)
    monkeypatch.setenv("LIMEN_DATASET_PATH", str(tmp_path))
    status = discover_official_dataset()
    assert status.available is True
    assert status.resolved_root == str(tmp_path.resolve())


def test_dry_run_report_structure(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _write_mini_official_xlsx(tmp_path)
    monkeypatch.setenv("LIMEN_DATASET_PATH", str(tmp_path))
    from evals.llm.official_benchmark import run_official_dry_run

    report = run_official_dry_run()
    assert "ready_for_official_benchmark" in report
    assert report["sheet_required"] == RESULT_SHEET
    assert json.dumps(report)  # serializable
