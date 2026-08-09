"""Read-only official dataset xlsx loading (sheet `result` only)."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from evals.llm.official_dataset import PREFERRED_FILENAMES

RESULT_SHEET = "result"

# Explicit per-file field classification (never guess at prompt time).
DATASET_FINAL_FIELDS: dict[str, str] = {
    "dialogo_id": "METADATA",
    "caso_id": "METADATA",
    "paciente_id": "METADATA",
    "dia_postop": "METADATA",
    "turno_idx": "METADATA",
    "hablante": "MODEL_INPUT",
    "texto": "MODEL_INPUT",
    "label_ground_truth": "EVALUATION_ONLY",
    "estilo_paciente": "METADATA",
    "modelo_paciente": "METADATA",
    "modelo_agente": "METADATA",
    "capa": "METADATA",
    "generado_ts": "METADATA",
}

TRAYECTORIAS_FIELDS: dict[str, str] = {
    "trayectoria_id": "METADATA",
    "paciente_id": "METADATA",
    "dia_postop": "EVALUATION_ONLY",
    "arquetipo_trayectoria": "EVALUATION_ONLY",
    "dolor_nrs": "EVALUATION_ONLY",
    "fiebre_c": "EVALUATION_ONLY",
    "movilidad": "EVALUATION_ONLY",
    "herida": "EVALUATION_ONLY",
    "apetito": "EVALUATION_ONLY",
    "sueno": "EVALUATION_ONLY",
    "seed": "METADATA",
    "generado_ts": "METADATA",
}

PERFILES_CLINICOS_FIELDS: dict[str, str] = {
    "paciente_id": "METADATA",
    "bundle_id": "EVALUATION_ONLY",
    "synthea_runtime": "EVALUATION_ONLY",
    "modulo_synthea": "EVALUATION_ONLY",
    "procedimiento": "MODEL_ALLOWED",
    "fecha_cirugia": "MODEL_ALLOWED",
    "edad": "MODEL_ALLOWED",
    "genero": "MODEL_ALLOWED",
    "comorbilidades": "MODEL_ALLOWED",
    "complicacion_encounter": "EVALUATION_ONLY",
    "generado_ts": "METADATA",
}

PERFILES_PACIENTES_CO_FIELDS: dict[str, str] = {
    "paciente_id": "METADATA",
    "nombre_completo": "PII_METADATA",
    "direccion": "PII_METADATA",
    "ciudad": "PII_METADATA",
    "departamento": "PII_METADATA",
    "documento_cc": "PII_METADATA",
    "eps": "PII_METADATA",
    "source_country": "METADATA",
    "adapted_country": "METADATA",
    "adaptation_fields": "METADATA",
    "adaptation_ts": "METADATA",
}

MODEL_ALLOWED_CLINICAL_FIELDS: frozenset[str] = frozenset(
    k for k, v in PERFILES_CLINICOS_FIELDS.items() if v == "MODEL_ALLOWED"
)


@dataclass(frozen=True)
class OfficialDatasetPaths:
    root: Path
    dataset_final: Path
    trayectorias: Path
    perfiles_clinicos: Path
    perfiles_pacientes_co: Path


@dataclass
class LoadedOfficialTables:
    turns: list[dict[str, Any]]
    trajectories: list[dict[str, Any]]
    clinical_profiles: list[dict[str, Any]]
    patient_demographics: list[dict[str, Any]]
    field_classification: dict[str, dict[str, str]]


def resolve_official_paths(root: Path) -> OfficialDatasetPaths | None:
    """Resolve the four canonical xlsx files under a dataset root."""
    if not root.exists():
        return None
    resolved: dict[str, Path] = {}
    for name in PREFERRED_FILENAMES:
        candidate = root / name if root.is_dir() else root
        if root.is_file() and root.name == name:
            candidate = root
        elif root.is_dir():
            candidate = root / name
        if candidate.is_file():
            resolved[name] = candidate.resolve()
    if len(resolved) != len(PREFERRED_FILENAMES):
        return None
    return OfficialDatasetPaths(
        root=root.resolve() if root.is_dir() else root.parent.resolve(),
        dataset_final=resolved["dataset_final.xlsx"],
        trayectorias=resolved["trayectorias_postop_silver.xlsx"],
        perfiles_clinicos=resolved["perfiles_clinicos_pacientes_silver_contest.xlsx"],
        perfiles_pacientes_co=resolved["perfiles_pacientes_co.xlsx"],
    )


def _load_xlsx_rows(path: Path, *, sheet: str = RESULT_SHEET) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError("openpyxl required for official dataset loading") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"sheet {sheet!r} missing in {path.name}; found {wb.sheetnames}")
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return []
        columns = [str(c).strip() if c is not None else "" for c in header_row]
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            record: dict[str, Any] = {}
            for idx, col in enumerate(columns):
                if not col:
                    continue
                value = row[idx] if idx < len(row) else None
                record[col] = value
            out.append(record)
        return out
    finally:
        wb.close()


def load_official_tables(paths: OfficialDatasetPaths) -> LoadedOfficialTables:
    """Load all official tables from xlsx (read-only, sheet `result`)."""
    turns = _load_xlsx_rows(paths.dataset_final)
    trajectories = _load_xlsx_rows(paths.trayectorias)
    clinical_profiles = _load_xlsx_rows(paths.perfiles_clinicos)
    patient_demographics = _load_xlsx_rows(paths.perfiles_pacientes_co)
    return LoadedOfficialTables(
        turns=turns,
        trajectories=trajectories,
        clinical_profiles=clinical_profiles,
        patient_demographics=patient_demographics,
        field_classification={
            "dataset_final.xlsx": dict(DATASET_FINAL_FIELDS),
            "trayectorias_postop_silver.xlsx": dict(TRAYECTORIAS_FIELDS),
            "perfiles_clinicos_pacientes_silver_contest.xlsx": dict(PERFILES_CLINICOS_FIELDS),
            "perfiles_pacientes_co.xlsx": dict(PERFILES_PACIENTES_CO_FIELDS),
        },
    )


def classify_table_columns(filename: str, columns: list[str]) -> dict[str, list[str]]:
    """Return explicit classification buckets for a table."""
    schema = {
        "dataset_final.xlsx": DATASET_FINAL_FIELDS,
        "trayectorias_postop_silver.xlsx": TRAYECTORIAS_FIELDS,
        "perfiles_clinicos_pacientes_silver_contest.xlsx": PERFILES_CLINICOS_FIELDS,
        "perfiles_pacientes_co.xlsx": PERFILES_PACIENTES_CO_FIELDS,
    }.get(filename, {})
    buckets: dict[str, list[str]] = {
        "MODEL_INPUT": [],
        "MODEL_ALLOWED": [],
        "EVALUATION_ONLY": [],
        "METADATA": [],
        "PII_METADATA": [],
        "UNKNOWN": [],
    }
    for col in columns:
        kind = schema.get(col, "UNKNOWN")
        buckets.setdefault(kind, []).append(col)
    return buckets
