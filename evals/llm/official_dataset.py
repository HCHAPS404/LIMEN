"""Official challenge dataset discovery + read-only loader (PHASE 5C).

Never fabricates files. Never alters spreadsheets.
Ground truth is EVALUATION_ONLY — never serialized into model prompts.

Canonical resolution order (no home-directory recursion):

1. LIMEN_DATASET_PATH
2. ./dataset/
3. ./data/challenge/
4. unavailable
"""

from __future__ import annotations

import hashlib
import os
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]

# Exact filenames preferred for challenge materials (non-recursive home walks).
PREFERRED_FILENAMES: tuple[str, ...] = (
    "dataset_final.xlsx",
    "trayectorias_postop_silver.xlsx",
    "perfiles_clinicos_pacientes_silver_contest.xlsx",
    "perfiles_pacientes_co.xlsx",
)

# Shallow extra globs inside a resolved root only (not $HOME).
ROOT_GLOBS: tuple[str, ...] = (
    "*.xlsx",
    "*.csv",
    "capa1*.xlsx",
    "capa2*.xlsx",
    "*trajectory*.xlsx",
    "*trayectorias*.xlsx",
    "*perfiles*.xlsx",
    "*dataset*.xlsx",
    "*dataset*.csv",
)

PROHIBITED_GROUND_TRUTH_FIELDS: frozenset[str] = frozenset(
    {
        "label_ground_truth",
        "expected_risk",
        "expected_answer",
        "expected_severity",
        "trajectory_outcome",
        "hidden_annotation",
        "challenge_label",
        "evaluator_notes",
        "ground_truth",
        "silver_label",
        "gold_label",
        "caso_truth",
        "label",
        "silver",
        "outcome",
        # Official trajectory / silver evaluation-only fields (PHASE 5C.2)
        "arquetipo_trayectoria",
        "dolor_nrs",
        "fiebre_c",
        "complicacion_encounter",
        "trayectoria_id",
    }
)

# Trajectory silver fields — block field-name tokens, not bare Spanish words like "herida".
TRAJECTORY_FIELD_TOKENS: frozenset[str] = frozenset(
    {
        "dolor_nrs",
        "fiebre_c",
        "arquetipo_trayectoria",
        "complicacion_encounter",
        "trayectoria_id",
    }
)

PROHIBITED_PROMPT_SUBSTRINGS: frozenset[str] = frozenset(
    {
        "label_ground_truth",
        "arquetipo_trayectoria",
        "dolor_nrs",
        "fiebre_c",
        "complicacion_encounter",
        "trayectoria_id",
    }
)

MODEL_INPUT_FIELD_HINTS: frozenset[str] = frozenset(
    {
        "patient_text",
        "utterance",
        "transcript",
        "texto",
        "mensaje",
        "turn_text",
        "user_text",
        "clinical_text",
        "texto_clinico",
    }
)

EVALUATION_ONLY_FIELD_HINTS: frozenset[str] = frozenset(
    {
        "label_ground_truth",
        "expected_risk",
        "expected_answer",
        "expected_severity",
        "trajectory_outcome",
        "silver_label",
        "gold_label",
        "ground_truth",
        "challenge_label",
    }
)

METADATA_FIELD_HINTS: frozenset[str] = frozenset(
    {
        "caso_id",
        "case_id",
        "patient_id",
        "split",
        "layer",
        "capa",
        "source_file",
    }
)


@dataclass
class DatasetFileFingerprint:
    filename: str
    path: str
    sha256: str
    row_count: int | str
    columns: list[str] | str
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class OfficialDatasetStatus:
    available: bool
    status: str  # AVAILABLE | UNAVAILABLE
    resolution_order: list[str] = field(default_factory=list)
    resolved_root: str | None = None
    files_found: list[str] = field(default_factory=list)
    fingerprints: list[dict[str, Any]] = field(default_factory=list)
    evaluation_enabled: bool = False
    clean_noisy_supported: bool = False
    dataset_path_env: str | None = None
    operator_instructions: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "OFFICIAL_DATASET": self.status,
            "available": self.available,
            "resolution_order": self.resolution_order,
            "resolved_root": self.resolved_root,
            "files_found": self.files_found,
            "fingerprints": self.fingerprints,
            "evaluation_enabled": self.evaluation_enabled,
            "clean_noisy_supported": self.clean_noisy_supported,
            "dataset_path_env": self.dataset_path_env,
            "operator_instructions": self.operator_instructions,
            "notes": self.notes,
        }


def canonical_resolution_candidates(project_root: Path | None = None) -> list[tuple[str, Path]]:
    """Prefer this explicit dataset resolution order (no $HOME recursion)."""
    base = project_root or ROOT
    ordered: list[tuple[str, Path]] = []
    env = (os.environ.get("LIMEN_DATASET_PATH") or "").strip()
    if env:
        ordered.append(("LIMEN_DATASET_PATH", Path(env).expanduser().resolve()))
    ordered.append(("./dataset/", (base / "dataset").resolve()))
    ordered.append(("./data/challenge/", (base / "data" / "challenge").resolve()))
    return ordered


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _inspect_tabular(path: Path) -> tuple[int | str, list[str] | str, list[str]]:
    """Best-effort row/column inspect. Never mutates the file."""
    notes: list[str] = []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        try:
            import csv

            with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            if not rows:
                return 0, [], notes
            columns = [str(c) for c in rows[0]]
            return max(0, len(rows) - 1), columns, notes
        except Exception as exc:
            notes.append(f"csv_inspect_failed:{type(exc).__name__}")
            return "UNMEASURED", "UNMEASURED", notes

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            notes.append("openpyxl_unavailable; columns/rows UNMEASURED")
            return "UNMEASURED", "UNMEASURED", notes
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                wb.close()
                return 0, [], notes
            columns = [str(c) if c is not None else "" for c in header]
            count = 0
            for _ in rows_iter:
                count += 1
            wb.close()
            return count, columns, notes
        except Exception as exc:
            notes.append(f"xlsx_inspect_failed:{type(exc).__name__}")
            return "UNMEASURED", "UNMEASURED", notes

    notes.append(f"unsupported_extension:{suffix or 'none'}")
    return "UNMEASURED", "UNMEASURED", notes


def fingerprint_dataset_file(path: Path) -> DatasetFileFingerprint:
    row_count, columns, notes = _inspect_tabular(path)
    try:
        sha = _sha256_file(path)
    except OSError as exc:
        sha = "UNMEASURED"
        notes.append(f"sha256_failed:{type(exc).__name__}")
    return DatasetFileFingerprint(
        filename=path.name,
        path=str(path.resolve()),
        sha256=sha,
        row_count=row_count,
        columns=columns,
        notes=notes,
    )


def _collect_files_in_root(root: Path) -> list[Path]:
    if not root.exists():
        return []
    if root.is_file():
        return [root]
    found: list[Path] = []
    for name in PREFERRED_FILENAMES:
        candidate = root / name
        if candidate.is_file():
            found.append(candidate)
    # Also accept a texts/ clinical directory marker (no deep home walk).
    for pattern in ROOT_GLOBS:
        for path in root.glob(pattern):
            if path.is_file() and path not in found:
                found.append(path)
        for path in (root / "texts").glob(pattern) if (root / "texts").is_dir() else []:
            if path.is_file() and path not in found:
                found.append(path)
    return sorted(found, key=lambda p: p.name.lower())


def discover_official_dataset(project_root: Path | None = None) -> OfficialDatasetStatus:
    """Resolve official dataset via canonical order only."""
    base = project_root or ROOT
    env = (os.environ.get("LIMEN_DATASET_PATH") or "").strip() or None
    order = canonical_resolution_candidates(base)
    resolution_labels = [label for label, _ in order] + ["unavailable"]

    for label, root in order:
        files = _collect_files_in_root(root)
        if not files:
            continue
        fingerprints = [fingerprint_dataset_file(path).to_dict() for path in files]
        preferred_set = {root / name for name in PREFERRED_FILENAMES}
        has_all_preferred = preferred_set.issubset({p.resolve() for p in files})
        evaluation_enabled = False
        clean_noisy_supported = False
        notes = [
            f"Resolved via {label}.",
            "Do not use ground-truth fields in model prompts.",
            "Fingerprints are metadata only; datasets are not committed.",
        ]
        if has_all_preferred:
            notes.append(
                "All four canonical xlsx files present. "
                "Run `make verify-official-dataset` to validate schema and reconstruction."
            )
        else:
            notes.append("Partial file set; official evaluation remains disabled.")
        return OfficialDatasetStatus(
            available=True,
            status="AVAILABLE",
            resolution_order=resolution_labels,
            resolved_root=str(root),
            files_found=[str(p.resolve()) for p in files],
            fingerprints=fingerprints,
            dataset_path_env=env,
            evaluation_enabled=evaluation_enabled,
            clean_noisy_supported=clean_noisy_supported,
            notes=notes,
        )

    return OfficialDatasetStatus(
        available=False,
        status="UNAVAILABLE",
        resolution_order=resolution_labels,
        resolved_root=None,
        files_found=[],
        fingerprints=[],
        dataset_path_env=env,
        operator_instructions=[
            "Official Tech Sphere dataset UNAVAILABLE under canonical resolution.",
            "Prefer this explicit order:",
            "  1. export LIMEN_DATASET_PATH=/absolute/path/to/official/dataset",
            "  2. ./dataset/",
            "  3. ./data/challenge/",
            "  4. unavailable",
            "Expected filenames (examples):",
            "  dataset_final.xlsx",
            "  trayectorias_postop_silver.xlsx",
            "  perfiles_clinicos_pacientes_silver_contest.xlsx",
            "  perfiles_pacientes_co.xlsx",
            "Do NOT recursively scan the home directory.",
            "Do not download unofficial copies. Do not commit labeled datasets.",
        ],
        notes=[
            "Synthetic control benchmark proceeds without fabricating official metrics.",
        ],
    )


def classify_columns(columns: list[str]) -> dict[str, list[str]]:
    """Heuristic column classification for read-only schema inspect."""
    model_input: list[str] = []
    evaluation_only: list[str] = []
    metadata: list[str] = []
    unknown: list[str] = []
    for col in columns:
        cl = col.strip().lower()
        if cl in EVALUATION_ONLY_FIELD_HINTS or any(
            h in cl for h in ("ground_truth", "silver", "label_gt", "expected_")
        ):
            evaluation_only.append(col)
        elif cl in MODEL_INPUT_FIELD_HINTS or any(
            h in cl for h in ("texto", "utterance", "transcript", "mensaje")
        ):
            model_input.append(col)
        elif cl in METADATA_FIELD_HINTS or cl.endswith("_id"):
            metadata.append(col)
        else:
            unknown.append(col)
    return {
        "MODEL_INPUT_FIELDS": model_input,
        "EVALUATION_ONLY_FIELDS": evaluation_only,
        "METADATA_FIELDS": metadata,
        "UNKNOWN_FIELDS": unknown,
    }


def filter_model_input_fields(record: dict[str, Any]) -> dict[str, Any]:
    """Strip prohibited ground-truth fields before building prompts."""
    out: dict[str, Any] = {}
    for k, v in record.items():
        kl = k.lower()
        if kl in PROHIBITED_GROUND_TRUTH_FIELDS:
            continue
        if any(
            p in kl
            for p in (
                "ground_truth",
                "silver_label",
                "expected_risk",
                "expected_answer",
            )
        ):
            continue
        out[k] = v
    return out


def assert_no_ground_truth_in_prompt(prompt: str, record: dict[str, Any] | None = None) -> None:
    """Runtime firewall: fail closed if prohibited GT leaks into prompts."""
    lowered = prompt.lower()
    for substring in PROHIBITED_PROMPT_SUBSTRINGS:
        if substring in lowered:
            raise AssertionError(f"prohibited ground-truth field leaked into prompt: {substring}")
    for gt_field in PROHIBITED_GROUND_TRUTH_FIELDS:
        if gt_field in TRAJECTORY_FIELD_TOKENS:
            continue
        if gt_field in {"label", "silver", "outcome"}:
            if f"{gt_field}=" in lowered or f"{gt_field}:" in lowered:
                raise AssertionError(
                    f"prohibited ground-truth field leaked into prompt: {gt_field}"
                )
            continue
        if gt_field in lowered:
            raise AssertionError(f"prohibited ground-truth field leaked into prompt: {gt_field}")
    for gt_field in TRAJECTORY_FIELD_TOKENS:
        if gt_field in lowered:
            raise AssertionError(f"prohibited trajectory field leaked into prompt: {gt_field}")
    for gt_field in TRAJECTORY_FIELD_TOKENS:
        if gt_field in lowered:
            raise AssertionError(f"prohibited trajectory field leaked into prompt: {gt_field}")
    if record:
        for gt_field in PROHIBITED_GROUND_TRUTH_FIELDS:
            keys_l = {k.lower(): k for k in record}
            if gt_field not in keys_l:
                continue
            val = record.get(keys_l[gt_field])
            if val is not None and str(val) and str(val) in prompt:
                raise AssertionError(f"ground-truth value for {gt_field} leaked into prompt")


def firewall_prompt(prompt: str, *, purpose: str = "benchmark") -> str:
    """Assert + return prompt unchanged (side-effect: raise on leak)."""
    assert_no_ground_truth_in_prompt(prompt)
    _ = purpose
    return prompt


@dataclass
class OfficialCase:
    """Read-only official case. Ground truth held separately from model_input."""

    case_id: str
    layer: str  # clean | noisy | unknown
    model_input: dict[str, Any]
    evaluation_only: dict[str, Any] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)


def load_official_cases_readonly(_paths: list[str]) -> list[OfficialCase]:
    """Placeholder loader — enabled only after schema confirmation.

    Returns empty until evaluation_enabled is flipped after inspect.
    Never mutates source files.
    """
    return []
